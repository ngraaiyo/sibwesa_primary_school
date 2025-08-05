
# students/views.py

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required, user_passes_test
from .models import Student, Class, Subject, Examination, Mark # 
from django.forms import modelformset_factory, inlineformset_factory
from .forms import ( StudentForm, 
                    ClassForm, 
                    SubjectForm, 
                    ExaminationForm, 
                    MarkEntrySelectionForm, 
                    MarkExcelUploadForm, 
                    ResultSelectionForm, 
                    StudentCreationForm
        )

from django import forms
import datetime
from datetime import date
from .forms import StudentExcelUploadForm
from django.urls import reverse
from django.contrib.admin.views.decorators import staff_member_required
from django.views.generic import ListView, CreateView, UpdateView, DeleteView
from django.urls import reverse_lazy
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from .models import SchoolDocument
from .forms import SchoolDocumentForm
from django.contrib.auth.models import User

from django.http import HttpResponse
from django.template.loader import render_to_string
from weasyprint import HTML
import tempfile

from django.db.models import Q
import pandas as pd
import openpyxl
from users.models import CustomUser
from students.models import Student, Class
from django.core.exceptions import ObjectDoesNotExist
from decimal import Decimal

def is_admin(user):
    return user.is_authenticated and user.role == 'admin'

def is_any_teacher(user):
    return user.is_authenticated and user.role in [
        'headteacher',
        'class_teacher',
        'academic_teacher',
        'statistic_teacher',
        'subject_teacher', 
    ]

def is_headteacher(user): 
    return user.is_authenticated and user.role == 'headteacher'

def get_teacher_assigned_classes(user):

    try:
        if user.role == 'class_teacher' and hasattr(user, 'teacher_profile') and user.teacher_profile.class_assigned:
            return Class.objects.filter(pk=user.teacher_profile.class_assigned.pk)
    except AttributeError:
        pass # User might not have a teacher_profile or class_assigned

    return Class.objects.none() # Return an empty queryset if no class is assigned

def is_academic_teacher(user):
    return user.is_authenticated and user.role == 'academic_teacher'

def is_class_teacher(user):
    return user.is_authenticated and user.role == 'class_teacher'
    
def is_subject_teacher(user):
    return user.is_authenticated and user.role == 'subject_teacher'

def can_view_all_students_and_add(user): 
    return is_admin(user) or is_headteacher(user) or is_academic_teacher(user) or is_statistic_teacher(user)

def is_statistic_teacher(user):  
    return user.is_authenticated and user.role == 'statistic_teacher'

def is_admin_or_academic_teacher(user):
    return is_admin(user) or is_academic_teacher(user)

def is_admin_or_headteacher_or_statistic_teacher(user):
    return is_admin(user) or is_headteacher(user) or is_statistic_teacher(user)

def is_admin_or_teacher(user):
    return is_admin(user) or is_headteacher(user) or is_academic_teacher(user) or is_class_teacher(user) or is_statistic_teacher(user)

def is_admin_or_headteacher(user):
    return is_admin(user) or is_headteacher(user)

def can_access_all_students(user):
    return is_admin(user) or is_academic_teacher(user)

def can_access_my_class_students(user):
    return is_class_teacher(user)

def is_general_school_dashboard_user(user):
    return user.role in [
        'class_teacher',
        'academic_teacher',
        'headteacher',
        'statistic_teacher',
        'subject_teacher',
    ]

def is_teacher(user):
     return user.is_authenticated and Class.objects.filter(class_teacher=user).exists()

def get_grade(score):
    if score is None:
        return "N/A"
    if score >= 81:
        return "A"
    elif score >= 61:
        return "B"
    elif score >= 41:
        return "C"
    elif score >= 21:
        return "D"
    else:
        return "F"

def calculate_results(examination, class_obj):
    
    students_in_class = Student.objects.filter(current_class=class_obj).order_by('first_name', 'last_name')

    all_subjects_for_class = Subject.objects.filter(
     
    ).distinct().order_by('name') # Get all subjects for consistent display

    class_results = []
    student_total_scores = {} # To store total scores for positioning

    for student in students_in_class:
        # Get all marks for the current student in the given examination
        student_marks = Mark.objects.filter(
            student=student,
            examination=examination
        ).select_related('subject')

        subject_details = []
        current_student_total_score = 0
        current_student_subject_count = 0 # To count subjects with actual scores

        # Iterate through all relevant subjects for this class/exam
        for subject in all_subjects_for_class:
            mark_for_subject = student_marks.filter(subject=subject).first() # Get the specific mark for this subject

            score = mark_for_subject.score if mark_for_subject else None
            grade = get_grade(score) if score is not None else "N/A" # Handle N/A for grade too

            subject_details.append({
                'subject_name': subject.name,
                'subject_code': subject.code,
                'score': score,
                'grade': grade,
            })

            if score is not None:
                current_student_total_score += score
                current_student_subject_count += 1

        average_score = current_student_total_score / current_student_subject_count if current_student_subject_count > 0 else 0
        overall_grade = get_overall_grade(average_score) # You need a get_overall_grade function

        student_total_scores[student.pk] = current_student_total_score

        class_results.append({
            'student': student,
            'total_score': current_student_total_score,
            'average_score': round(average_score, 2),
            'overall_grade': overall_grade,
            'subject_details': subject_details, # This now includes all subjects
            'position': None, # Will be set later
        })

    # Sort students by total score to determine positions
    class_results_sorted = sorted(class_results, key=lambda x: x['total_score'], reverse=True)

    # Assign positions
    if class_results_sorted:
        class_results_sorted[0]['position'] = 1
        for i in range(1, len(class_results_sorted)):
            if class_results_sorted[i]['total_score'] == class_results_sorted[i-1]['total_score']:
                class_results_sorted[i]['position'] = class_results_sorted[i-1]['position']
            else:
                class_results_sorted[i]['position'] = i + 1

    return class_results_sorted

# You'll also need to ensure you have a get_overall_grade function
def get_overall_grade(average_score):
    if average_score >= 81:
        return 'A'
    elif average_score >= 61:
        return 'B'
    elif average_score >= 41:
        return 'C'
    elif average_score >= 21:
        return 'D'
    else:
        return 'F'

@login_required
@user_passes_test(can_access_all_students)
def student_list(request):
    all_students = Student.objects.all().order_by('current_class__name', 'first_name')
    class_students = None # For class teacher's specific class students

    if is_admin(request.user) or is_academic_teacher(request.user):
        # Admin and Academic Teachers see all students
        context = {
            'all_students': all_students,
            'can_add_edit_delete': is_admin(request.user) # Only admin can add/edit/delete
        }
    elif is_class_teacher(request.user):
        # Class Teacher sees their assigned class students on one side and all students on another
        try:
            assigned_class = Class.objects.get(class_teacher=request.user)
            class_students = Student.objects.filter(current_class=assigned_class).order_by('first_name')
            context = {
                'all_students': all_students, # All students on one side
                'class_students': class_students, # Class-specific students on the other
                'assigned_class_name': assigned_class.name,
                'can_add_edit_delete': False # Class Teacher cannot add/edit/delete students
            }
        except Class.DoesNotExist:
            messages.warning(request, "You are a Class Teacher but not assigned to any class yet.")
            context = {
                'all_students': all_students,
                'class_students': None,
                'can_add_edit_delete': False
            }
    else:
        # This case should ideally be caught by user_passes_test, but as a fallback
        messages.error(request, "You are not authorized to view student records.")
        return redirect('teacher_dashboard') # Redirect to general dashboard or login

    return render(request, 'students/student_list.html', context)

@login_required
@user_passes_test(is_admin_or_headteacher_or_statistic_teacher, login_url='/users/login/') 
def student_add(request):
    if request.method == 'POST':
        form = StudentForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Student added successfully!')
            return redirect('all_students')
    else:
        form = StudentForm()
    return render(request, 'students/student_form.html', {'form': form, 'title': 'Add New Student'})

@login_required
@user_passes_test(is_admin_or_headteacher, login_url='/users/login/') 
def student_edit(request, pk):
    student = get_object_or_404(Student, pk=pk)
    if request.method == 'POST':
        form = StudentForm(request.POST, instance=student)
        if form.is_valid():
            form.save()
            messages.success(request, 'Student updated successfully!')
            return redirect('student_list')
    else:
        form = StudentForm(instance=student)
    return render(request, 'students/student_form.html', {'form': form, 'title': 'Edit Student'})

@login_required
@user_passes_test(is_admin_or_headteacher, login_url='/users/login/') # Only Admin can delete students
def student_delete(request, pk):
    student = get_object_or_404(Student, pk=pk)
    if request.method == 'POST':
        student.delete()
        messages.success(request, 'Student deleted successfully!')
        return redirect('all_students')
    return render(request, 'students/student_confirm_delete.html', {'student': student})

@login_required
@user_passes_test(is_admin_or_headteacher, login_url='/users/login/') # Only Admin can upload Excel
def student_upload_excel(request):
    if request.method == 'POST':
        excel_file = request.FILES.get('excel_file')
        if not excel_file:
            messages.error(request, 'No file uploaded.')
            return render(request, 'students/student_upload_excel.html')

        if not excel_file.name.endswith(('.xlsx', '.xls')):
            messages.error(request, 'Invalid file type. Please upload an Excel file (.xlsx or .xls).')
            return render(request, 'students/student_upload_excel.html')

        try:
            workbook = openpyxl.load_workbook(excel_file)
            sheet = workbook.active
            header = [cell.value for cell in sheet[1]] # Get header row

            # Expected headers (case-insensitive)
            expected_headers = {
                'first name': 'first_name',
                'middle name': 'middle_name',
                'last name': 'last_name',
                'date of birth': 'date_of_birth',
                'gender': 'gender',
                'admission number': 'admission_number',
                'class name': 'current_class_name', # We'll map this to actual Class object
                'class year': 'current_class_year', # Used with class name
            }

            # Map actual headers to expected model fields
            header_map = {}
            for col_idx, h in enumerate(header):
                if h and str(h).strip().lower() in expected_headers:
                    header_map[expected_headers[str(h).strip().lower()]] = col_idx

            required_headers = ['first_name', 'last_name', 'date_of_birth', 'gender', 'admission_number', 'current_class_name', 'current_class_year']
            if not all(rh in header_map for rh in required_headers):
                missing_headers = [h.replace('_', ' ').title() for h in required_headers if h not in header_map]
                messages.error(request, f"Missing required headers in Excel: {', '.join(missing_headers)}")
                return render(request, 'students/student_upload_excel.html')

            students_added = 0
            errors = []

            for row_idx, row in enumerate(sheet.iter_rows(min_row=2), start=2): # Start from second row
                row_data = {field_name: row[col_idx].value for field_name, col_idx in header_map.items()}

                try:
                    # Validate and clean data
                    first_name = row_data.get('first_name')
                    middle_name = row_data.get('middle_name')
                    last_name = row_data.get('last_name')
                    date_of_birth_raw = row_data.get('date_of_birth') # Get raw value
                    gender = str(row_data.get('gender')).strip().upper()
                    admission_number = str(row_data.get('admission_number')).strip()
                    class_name = str(row_data.get('current_class_name')).strip()

                    # Handle class_year potentially being None or non-integer
                    class_year_raw = row_data.get('current_class_year')
                    if class_year_raw is None or str(class_year_raw).strip() == '':
                        class_year = None
                    else:
                        try:
                            class_year = int(class_year_raw)
                        except (ValueError, TypeError):
                            errors.append(f"Row {row_idx}: Invalid Class Year ('{class_year_raw}'). Must be an integer.")
                            continue # Skip this row

                    # Basic validation for non-date fields
                    if not all([first_name, last_name, gender, admission_number, class_name, class_year is not None]):
                        errors.append(f"Row {row_idx}: Missing required data (first name, last name, gender, admission number, class name, or class year).")
                        continue

                    # --- START NEW ROBUST DATE PARSING LOGIC ---
                    date_of_birth = None
                    if isinstance(date_of_birth_raw, datetime.datetime):
                        date_of_birth = date_of_birth_raw.date() # Extract date part from datetime object
                    elif isinstance(date_of_birth_raw, str) and date_of_birth_raw.strip():
                        date_of_birth_str = date_of_birth_raw.strip().split(' ')[0] # Remove time if present

                        formats_to_try = [
                            '%Y-%m-%d',  # YYYY-MM-DD (e.g., 2000-01-15)
                            '%d-%m-%Y',  # DD-MM-YYYY (e.g., 15-01-2000)
                            '%m/%d/%Y',  # MM/DD/YYYY (e.g., 01/15/2000)
                            '%d/%m/%Y',  # DD/MM/YYYY (e.g., 15/01/2000)
                            '%Y/%m/%d',  # YYYY/MM/DD (e.g., 2000/01/15)
                            '%b %d, %Y',  # Jun 18, 2009 (Month Abbr Day, Year)
                            '%B %d, %Y',  # June 18, 2009 (Full Month Name Day, Year)
                            # Add other formats if you anticipate them, e.g., '%d %b %Y' for '18 Jun 2009'
                        ]

                        parsed = False
                        for fmt in formats_to_try:
                            try:
                                date_of_birth = datetime.datetime.strptime(date_of_birth_str, fmt).date()
                                parsed = True
                                break
                            except ValueError:
                                continue

                        if not parsed:
                            errors.append(f"Row {row_idx}: Invalid date format for Date of Birth ('{date_of_birth_str}'). Please use YYYY-MM-DD, DD-MM-YYYY, MM/DD/YYYY, or Month Day, Year format.")
                            continue # Skip row if date cannot be parsed
                    else: # If date_of_birth_raw is None or some other unexpected type
                        errors.append(f"Row {row_idx}: Date of Birth is missing or has an unexpected format/type ('{date_of_birth_raw}').")
                        continue

                    if not date_of_birth: # Final check after parsing attempts
                        errors.append(f"Row {row_idx}: Date of Birth is missing or could not be parsed.")
                        continue
                    # --- END NEW ROBUST DATE PARSING LOGIC ---

                    # Gender validation
                    if gender not in ['M', 'F', 'O']:
                        errors.append(f"Row {row_idx}: Invalid gender ('{gender}'). Use M, F, or O.")
                        continue

                    try:
                        student_class = Class.objects.get(name__iexact=class_name, year=class_year)
                    except Class.DoesNotExist:
                        errors.append(f"Row {row_idx}: Class '{class_name}' (Year {class_year}) not found. Please ensure the class exists.")
                        continue

                    student, created = Student.objects.update_or_create(
                        admission_number=admission_number,
                        defaults={
                            'first_name': first_name,
                            'middle_name': middle_name if middle_name else None,
                            'last_name': last_name,
                            'date_of_birth': date_of_birth,
                            'gender': gender,
                            'current_class': student_class
                        }
                    )
                    students_added += 1 if created else 0 # Count as added only if new, updated is implicit

                except Exception as e:
                    # This catches any other unexpected errors during row processing
                    errors.append(f"Row {row_idx}: Error processing row: {e}")
                    continue

            if students_added > 0:
                messages.success(request, f'Successfully uploaded! Added {students_added} students.')
            else:
                messages.info(request, "No new students were added (file might be empty or all students already exist/updated).")

            if errors:
                for error_msg in errors:
                    messages.error(request, error_msg)

            return redirect('student_list') # Redirect back to the student list

        except Exception as e:
            messages.error(request, f"An overall error occurred while processing the Excel file: {e}. Please check file format.")
            # print(f"DEBUG: Error processing Excel file: {e}") # For debugging
            return render(request, 'students/student_upload_excel.html')
    else:
        # This handles GET requests to display the upload form
        return render(request, 'students/student_upload_excel.html')

@login_required
@user_passes_test(is_admin_or_teacher, login_url='/users/login/')
def class_list(request):
    classes = Class.objects.all().order_by('year', 'name')

    # Only admin can add/edit/delete classes
    can_manage_classes = is_admin(request.user)

    context = {
        'classes': classes,
        'can_manage_classes': can_manage_classes
    }
    return render(request, 'students/class_list.html', context)

@login_required
@user_passes_test(is_admin, login_url='/users/login/')
def class_add(request):
    if request.method == 'POST':
        form = ClassForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Class added successfully!')
            return redirect('class_list')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field.replace('_', ' ').title()}: {error}")
    else:
        form = ClassForm()
    return render(request, 'students/class_form.html', {'form': form, 'title': 'Add New Class'})

@login_required
@user_passes_test(is_admin, login_url='/users/login/')
def class_edit(request, pk):
    class_obj = get_object_or_404(Class, pk=pk)
    if request.method == 'POST':
        form = ClassForm(request.POST, instance=class_obj)
        if form.is_valid():
            form.save()
            messages.success(request, 'Class updated successfully!')
            return redirect('class_list')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field.replace('_', ' ').title()}: {error}")
    else:
        form = ClassForm(instance=class_obj)
    return render(request, 'students/class_form.html', {'form': form, 'title': 'Edit Class'})

@login_required
@user_passes_test(is_admin, login_url='/users/login/')
def class_delete(request, pk):
    class_obj = get_object_or_404(Class, pk=pk)

    if class_obj.student_set.exists():
        messages.error(request, f"Cannot delete class '{class_obj.name} ({class_obj.year})' because there are students assigned to it. Please reassign or delete students first.")
        return redirect('class_list')

    if request.method == 'POST':
        class_obj.delete()
        messages.success(request, 'Class deleted successfully!')
        return redirect('class_list')
    return render(request, 'students/class_confirm_delete.html', {'class_obj': class_obj})

@login_required
@user_passes_test(is_admin_or_academic_teacher, login_url='/users/login/')
def subject_list(request):
    # This is a temporary placeholder. We will implement full logic later.
    messages.info(request, "Subject Management page is under construction!")
    return render(request, 'base.html', {'message': "Subject Management page is under construction!"})

@login_required
@user_passes_test(is_admin_or_academic_teacher, login_url='/users/login/')
def examination_list(request):
    # This is a temporary placeholder. We will implement full logic later.
    messages.info(request, "Examination Management page is under construction!")
    return render(request, 'base.html', {'message': "Examination Management page is under construction!"})

@login_required
@user_passes_test(is_admin_or_academic_teacher, login_url='/users/login/')
def subject_list(request):
    subjects = Subject.objects.all().order_by('name')
    can_manage_subjects = is_admin(request.user) 

    context = {
        'subjects': subjects,
        'can_manage_subjects': can_manage_subjects
    }
    return render(request, 'students/subject_list.html', context)

@login_required
@user_passes_test(is_admin, login_url='/users/login/')
def subject_add(request):
    if request.method == 'POST':
        form = SubjectForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Subject added successfully!')
            return redirect('subject_list')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field.replace('_', ' ').title()}: {error}")
    else:
        form = SubjectForm()
    return render(request, 'students/subject_form.html', {'form': form, 'title': 'Add New Subject'})

@login_required
@user_passes_test(is_admin, login_url='/users/login/')
def subject_edit(request, pk):
    subject = get_object_or_404(Subject, pk=pk)
    if request.method == 'POST':
        form = SubjectForm(request.POST, instance=subject)
        if form.is_valid():
            form.save()
            messages.success(request, 'Subject updated successfully!')
            return redirect('subject_list')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field.replace('_', ' ').title()}: {error}")
    else:
        form = SubjectForm(instance=subject)
    return render(request, 'students/subject_form.html', {'form': form, 'title': 'Edit Subject'})

@login_required
@user_passes_test(is_admin, login_url='/users/login/')
def subject_delete(request, pk):
    subject = get_object_or_404(Subject, pk=pk)

    if subject.examination_set.exists() or subject.mark_set.exists(): 
        messages.error(request, f"Cannot delete subject '{subject.name}' because it has associated examinations or marks.")
        return redirect('subject_list')

    if request.method == 'POST':
        subject.delete()
        messages.success(request, 'Subject deleted successfully!')
        return redirect('subject_list')
    return render(request, 'students/subject_confirm_delete.html', {'subject': subject})

@login_required
@user_passes_test(is_admin_or_academic_teacher, login_url='/users/login/')
def examination_list(request):
    examinations = Examination.objects.all().order_by('-academic_year', 'term', 'date', 'name')
    can_manage_examinations = is_admin(request.user) 

    context = {
        'examinations': examinations,
        'can_manage_examinations': can_manage_examinations
    }
    return render(request, 'students/examination_list.html', context)

@login_required
@user_passes_test(is_admin, login_url='/users/login/')
def examination_add(request):
    if request.method == 'POST':
        form = ExaminationForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Examination added successfully!')
            return redirect('examination_list')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field.replace('_', ' ').title()}: {error}")
    else:
        form = ExaminationForm()
    return render(request, 'students/examination_form.html', {'form': form, 'title': 'Add New Examination'})

@login_required
@user_passes_test(is_admin, login_url='/users/login/')
def examination_edit(request, pk):
    examination = get_object_or_404(Examination, pk=pk)
    if request.method == 'POST':
        form = ExaminationForm(request.POST, instance=examination)
        if form.is_valid():
            form.save()
            messages.success(request, 'Examination updated successfully!')
            return redirect('examination_list')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field.replace('_', ' ').title()}: {error}")
    else:
        form = ExaminationForm(instance=examination)
    return render(request, 'students/examination_form.html', {'form': form, 'title': 'Edit Examination'})

@login_required
@user_passes_test(is_admin, login_url='/users/login/')
def examination_delete(request, pk):
    examination = get_object_or_404(Examination, pk=pk)

    if examination.mark_set.exists():
        messages.error(request, f"Cannot delete examination '{examination.name}' because it has associated marks. Please delete marks for this examination first.")
        return redirect('examination_list')

    if request.method == 'POST':
        examination.delete()
        messages.success(request, 'Examination deleted successfully!')
        return redirect('examination_list')
    return render(request, 'students/examination_confirm_delete.html', {'examination': examination})

@login_required
@user_passes_test(is_general_school_dashboard_user, login_url='/users/login/')
def mark_entry_selection(request):
    if request.method == 'POST':
        form = MarkEntrySelectionForm(request.POST)

        if form.is_valid():
            examination = form.cleaned_data.get('examination')
            class_name = form.cleaned_data.get('class_name')
            subject = form.cleaned_data.get('subject')

            params = {}
            if examination:
                params['exam_id'] = examination.pk
            if subject:
                params['subject_id'] = subject.pk
            if class_name:
                params['class_id'] = class_name.pk

            if not (examination and subject and class_name):
                messages.error(request, "Please ensure you select an Examination, Subject, and Class.")
                return render(request, 'students/mark_entry_selection.html', {'form': form})

            return redirect(reverse('mark_entry_form', kwargs={
                'exam_id': examination.pk,
                'subject_id': subject.pk,
                'class_id': class_name.pk
            }))

        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field.replace('_', ' ').title()}: {error}")
            return render(request, 'students/mark_entry_selection.html', {'form': form})
    else: # GET request
        form = MarkEntrySelectionForm()

        # Class teacher pre-filling logic
        if is_class_teacher(request.user):
            # The class teacher can now choose from ALL classes
            form.fields['class_name'].queryset = Class.objects.all()
            form.fields['class_name'].required = True
            
            # As a helpful default, we'll try to pre-select their assigned class,
            # but allow them to change it.
            try:
                assigned_class = Class.objects.get(class_teacher=request.user)
                form.fields['class_name'].initial = assigned_class.pk
                messages.info(request, f"You are assigned to {assigned_class.name}. You can change this selection if needed.")
            except Class.DoesNotExist:
                # If they are a class teacher but not assigned to a class, we'll leave it blank
                messages.warning(request, "You are a Class Teacher but not assigned to any class. Please select a class.")
            except Class.MultipleObjectsReturned:
                messages.warning(request, "You are assigned to multiple classes. Please select one.")

            # For the subjects, we'll assume they can enter marks for all subjects they teach
            # This is a bit ambiguous, so we'll set it to all for simplicity. You can adjust this.
            form.fields['subject'].queryset = Subject.objects.all()
            form.fields['subject'].required = True


        elif is_academic_teacher(request.user):
            form.fields['class_name'].required = True # Academic teachers typically need to select a class
            form.fields['class_name'].queryset = Class.objects.all() # Allow them to pick any class
            form.fields['subject'].queryset = Subject.objects.filter(teachers=request.user)
            form.fields['subject'].required = True

        elif is_subject_teacher(request.user):
            form.fields['class_name'].queryset = Class.objects.all()
            form.fields['class_name'].required = True
            form.fields['subject'].queryset = Subject.objects.filter(teachers=request.user)
            form.fields['subject'].required = True

        elif is_headteacher(request.user) or is_statistic_teacher(request.user):
            form.fields['class_name'].queryset = Class.objects.all()
            form.fields['subject'].queryset = Subject.objects.all()
            form.fields['class_name'].required = True
            form.fields['subject'].required = True
        # --- END OF ADDED BLOCKS ---

    return render(request, 'students/mark_entry_selection.html', {'form': form})

@login_required
@user_passes_test(is_general_school_dashboard_user, login_url='/users/login/')
def mark_entry_form(request, exam_id, subject_id, class_id):
    print("User:", request.user)
    print("User authenticated?", request.user.is_authenticated)
    print("User role:", getattr(request.user, 'role', 'No role attribute'))

    try:
        examination = get_object_or_404(Examination, pk=exam_id)
        subject = get_object_or_404(Subject, pk=subject_id)
        selected_class = None
        if class_id:
            selected_class = get_object_or_404(Class, pk=class_id)
    except (Examination.DoesNotExist, Subject.DoesNotExist, Class.DoesNotExist):
        messages.error(request, "One of the selected Examination, Class, or Subject does not exist.")
        return redirect('mark_entry_selection')

    if is_class_teacher(request.user):
        try:
            user_assigned_class = Class.objects.get(class_teacher=request.user)
            if selected_class and selected_class.pk != user_assigned_class.pk:
                messages.error(request, "You can only enter marks for your assigned class.")
                return redirect('mark_entry_selection')
            students = Student.objects.filter(current_class=user_assigned_class).order_by('first_name', 'last_name')
            selected_class = user_assigned_class # Ensure it's the correct one
        except Class.DoesNotExist:
            messages.error(request, "You are a Class Teacher but not assigned to any class. Please contact the administrator.")
            return redirect('mark_entry_selection')
    elif any([
        is_academic_teacher(request.user),
        is_admin(request.user),
        is_statistic_teacher(request.user),
        is_headteacher(request.user)
    ]):
        if selected_class:
            students = Student.objects.filter(current_class=selected_class).order_by('first_name', 'last_name')
        else:
            students = Student.objects.all().order_by('first_name', 'last_name')
    else:
        messages.error(request, "You do not have permission to enter marks.")
        return redirect('login') 

    if not students.exists():
        messages.warning(request, "No students found for the selected criteria.")
        return redirect('mark_entry_selection')
    
    class IndividualMarkForm(forms.ModelForm):
        class Meta:
            model = Mark
            fields = ['score']
        sscore = forms.IntegerField(
            min_value=0,
            max_value=100,
            required=False,
            widget=forms.NumberInput()
        )

    MarkFormSet = modelformset_factory(Mark, form=IndividualMarkForm, extra=0, can_delete=False)
    initial_data = []
    for student in students:
        mark, created = Mark.objects.get_or_create(
            student=student,
            subject=subject,
            examination=examination,
            defaults={'score': None} 
        )
        initial_data.append({
            'id': mark.id,
            'student_name': f"{student.first_name} {student.middle_name} {student.last_name}",
            'score': mark.score,
            'student_id': student.id 
        })
    queryset = Mark.objects.filter(
        student__in=students,
        subject=subject,
        examination=examination
    )

    if request.method == 'POST':
        formset = MarkFormSet(request.POST, queryset=queryset)
        if formset.is_valid():
            for form in formset:
                if form.has_changed(): 
                    mark_instance = form.save(commit=False)
                    if not mark_instance.pk: 
                        student_id_from_form = form.cleaned_data.get('student_id') 
                        mark_instance.student = Student.objects.get(pk=student_id_from_form)
                        mark_instance.subject = subject
                        mark_instance.examination = examination
                    mark_instance.save()
            messages.success(request, 'Marks saved successfully!')
            return redirect('mark_entry_selection')
        else:
            messages.error(request, "Please correct the errors below.")
            for form in formset:
                if form.errors:
                    for field, errors in form.errors.items():
                        for error in errors:
                            messages.error(request, f"Error for Student (ID: {form.initial.get('student_id')}): {field.replace('_', ' ').title()}: {error}")
    else:
        mark_instances_for_formset = []
        for student in students:
            mark_instance, created = Mark.objects.get_or_create(
                student=student,
                subject=subject,
                examination=examination
            )
            mark_instances_for_formset.append(mark_instance)

        formset = MarkFormSet(queryset=Mark.objects.filter(pk__in=[m.pk for m in mark_instances_for_formset]))

    context = {
        'examination': examination,
        'selected_class': selected_class,
        'subject': subject,
        'formset': formset,
        'students': students 
    }
    return render(request, 'students/mark_entry_form.html', context)

@login_required
@user_passes_test(is_general_school_dashboard_user, login_url='/users/login/')
def mark_excel_upload(request):
    # Ensure examination context is passed for GET requests
    examination = None
    exam_id_get = request.GET.get('exam_id')
    if exam_id_get:
        examination = get_object_or_404(Examination, pk=exam_id_get)

    if request.method == 'POST':
        form = MarkExcelUploadForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES['excel_file']
            exam_id_post = request.POST.get('exam_id') # Get from POST for submissions

            if not exam_id_post:
                messages.error(request, "Examination not specified for upload. Please select an examination first.")
                return redirect('mark_entry_selection') # Or render with form and message

            examination = get_object_or_404(Examination, pk=exam_id_post)

            if not excel_file.name.endswith('.xlsx'):
                messages.error(request, 'Invalid file type. Please upload an Excel (.xlsx) file.')
                return render(request, 'students/mark_excel_upload.html', {'form': form, 'examination': examination})

            try:
                workbook = openpyxl.load_workbook(excel_file)
                sheet = workbook.active

                header = [cell.value for cell in sheet[1]]
                print(f"\n--- DEBUG START ---")
                print(f"--- DEBUG: Raw headers from Excel: {header} ---")

                _initial_required_headers = ['Admission_Number', 'First_Name', 'Middle_Name', 'Last_Name', 'Subject_Code', 'Score']
                required_headers_normalized = {h.replace(' ', '_').lower() for h in _initial_required_headers}
                normalized_header = {h.replace(' ', '_').lower() for h in header if h}

                print(f"--- DEBUG: Normalized headers (as seen by Django): {normalized_header} ---")
                print(f"--- DEBUG: Required headers (target - normalized): {required_headers_normalized} ---")

                if not required_headers_normalized.issubset(normalized_header):
                    missing_headers_display = set(_initial_required_headers) - set(h.replace('_', ' ').title() for h in normalized_header if h)
                    messages.error(request, f'Invalid Excel file format. Missing one or more required columns: {", ".join(sorted(missing_headers_display))}. Expected: {", ".join(_initial_required_headers)}.')
                    print(f"--- DEBUG: Missing headers (from normalized check): {required_headers_normalized - normalized_header} ---")
                    print(f"--- DEBUG END ---\n")
                    return render(request, 'students/mark_excel_upload.html', {'form': form, 'examination': examination})

                # Now, when getting column indices, use the lowercase/underscored names:
                header_map = {h.replace(' ', '_').lower(): i for i, h in enumerate(header) if h}
                admission_col = header_map.get('admission_number')
                first_name_col = header_map.get('first_name')
                middle_name_col = header_map.get('middle_name')
                last_name_col = header_map.get('last_name')
                subject_col = header_map.get('subject_code')
                score_col = header_map.get('score')

                # Ensure all columns are found before proceeding with row processing (redundant check, but good for clarity)
                if any(col is None for col in [admission_col, first_name_col, middle_name_col, last_name_col, subject_col, score_col]):
                    messages.error(request, "One or more required columns were not found in the Excel file.")
                    return render(request, 'students/mark_excel_upload.html', {'form': form, 'examination': examination})

                # --- START: Debugging variables for row processing ---
                print("\n--- DEBUG: Starting row processing ---")
                errors = []
                processed_count = 0
                skipped_count = 0
                # --- END: Debugging variables for row processing ---

                for row_idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
                    # --- START: Debugging for each row ---
                    print(f"\n--- DEBUG: Processing Row {row_idx} ---")
                    # --- END: Debugging for each row ---

                    # Check if row is entirely empty (optional, but good for robust handling)
                    if not any(cell.value for cell in row):
                        print(f"  INFO: Row {row_idx} is empty, skipping.")
                        skipped_count += 1
                        continue

                    admission_number = str(row[admission_col].value).strip() if row[admission_col].value is not None else ''
                    excel_first_name = str(row[first_name_col].value).strip() if row[first_name_col].value is not None else ''
                    excel_middle_name = str(row[middle_name_col].value).strip() if row[middle_name_col].value is not None else ''
                    excel_last_name = str(row[last_name_col].value).strip() if row[last_name_col].value is not None else ''
                    subject_code = str(row[subject_col].value).strip() if row[subject_col].value is not None else ''
                    score_value = row[score_col].value

                    # --- START: Debugging raw and cleaned values ---
                    print(f"  Raw values: Adm='{row[admission_col].value}', Sub='{row[subject_col].value}', Score='{row[score_col].value}'")
                    print(f"  Cleaned values: Admission='{admission_number}', Subject='{subject_code}', Score='{score_value}'")
                    # --- END: Debugging raw and cleaned values ---

                    if not admission_number or not subject_code:
                        errors.append(f"Row {row_idx}: Admission Number or Subject Code is empty. Skipping row.")
                        skipped_count += 1
                        print(f"  SKIPPED: Row {row_idx} - Missing Admission or Subject Code.") # Debug
                        continue

                    student = None
                    try:
                        student = Student.objects.get(admission_number=admission_number)
                        print(f"  SUCCESS: Row {row_idx} - Student found: ID={student.id}, Name={student.first_name} {student.last_name}, Class={student.current_class.name}") # Debug
                    except Student.DoesNotExist:
                        errors.append(f"Row {row_idx}: Student with admission number '{admission_number}' not found.")
                        skipped_count += 1
                        print(f"  ERROR: Row {row_idx} - Student not found for Admission Number: {admission_number}") # Debug
                        continue
                    except Exception as e: # Catch any other unexpected student lookup errors
                        errors.append(f"Row {row_idx}: Unexpected error finding student '{admission_number}': {e}")
                        skipped_count += 1
                        print(f"  CRITICAL ERROR: Row {row_idx} - Unexpected error finding student: {e}") # Debug
                        continue

                    subject = None
                    try:
                        subject = Subject.objects.get(code=subject_code)
                        print(f"  SUCCESS: Row {row_idx} - Subject found: ID={subject.id}, Name={subject.name}, Code={subject.code}") # Debug
                    except Subject.DoesNotExist:
                        errors.append(f"Row {row_idx}: Subject with code '{subject_code}' not found.")
                        skipped_count += 1
                        print(f"  ERROR: Row {row_idx} - Subject not found for Code: {subject_code}") # Debug
                        continue
                    except Exception as e: # Catch any other unexpected subject lookup errors
                        errors.append(f"Row {row_idx}: Unexpected error finding subject '{subject_code}': {e}")
                        skipped_count += 1
                        print(f"  CRITICAL ERROR: Row {row_idx} - Unexpected error finding subject: {e}") # Debug
                        continue

                    score = None
                    try:
                        # Ensure score_value is not None before attempting conversion
                        if score_value is None:
                            raise ValueError("Score is empty.")
                        
                        score = int(score_value)
                        if not (0 <= score <= 100):
                            errors.append(f"Row {row_idx}: Score for {admission_number} ({subject_code}) must be between 0 and 100. Found: {score_value}.")
                            skipped_count += 1
                            print(f"  ERROR: Row {row_idx} - Invalid score range: {score_value}") # Debug
                            continue
                        print(f"  SUCCESS: Row {row_idx} - Score parsed successfully: {score}") # Debug
                    except (ValueError, TypeError) as e:
                        errors.append(f"Row {row_idx}: Invalid score value for {admission_number} ({subject_code}). Found: '{score_value}'. Score must be a number (Error: {e}).")
                        skipped_count += 1
                        print(f"  ERROR: Row {row_idx} - Invalid score type/value: {score_value} (Error: {e})") # Debug
                        continue
                    except Exception as e: # Catch any other unexpected score parsing errors
                        errors.append(f"Row {row_idx}: Unexpected error parsing score '{score_value}': {e}")
                        skipped_count += 1
                        print(f"  CRITICAL ERROR: Row {row_idx} - Unexpected error parsing score: {e}") # Debug
                        continue


                    try:
                        mark, created = Mark.objects.update_or_create(
                            student=student,
                            subject=subject,
                            examination=examination,
                            defaults={'score': score}
                        )
                        processed_count += 1
                        print(f"  SUCCESS: Row {row_idx} - Mark {'created' if created else 'updated'}: Student={student.admission_number}, Subject={subject.code}, Score={score}") # Debug

                    except Exception as e:
                        errors.append(f"Row {row_idx}: An error occurred while saving mark for {admission_number} ({subject_code}): {e}")
                        skipped_count += 1
                        print(f"  CRITICAL ERROR: Row {row_idx} - Failed to save mark: {e}") # Debug
                        continue

                # --- START: Debugging after loop ---
                print(f"\n--- DEBUG: Row processing complete. Processed: {processed_count}, Skipped: {skipped_count} ---")
                print("--- DEBUG END ---\n")
                # --- END: Debugging after loop ---

                if errors:
                    # Use warning for individual row errors so upload message can still be seen
                    for err in errors:
                        messages.warning(request, err)

                if processed_count > 0:
                    messages.success(request, f"Successfully uploaded marks for {processed_count} students.")
                elif skipped_count > 0:
                    messages.info(request, f"No marks were successfully uploaded, but {skipped_count} rows were skipped due to issues. Please check warnings above.")
                else:
                    messages.info(request, "No marks were found or processed from the Excel file.")

                # Redirect to the selection page after processing is complete
                return redirect('mark_entry_selection')

            except Exception as e:
                # This catch-all also needs a print to catch unexpected errors during file parsing
                print(f"--- DEBUG: An unhandled error occurred during file processing: {e} ---")
                messages.error(request, f"An unhandled error occurred while processing the Excel file: {e}. Please check file format.")
                return render(request, 'students/mark_excel_upload.html', {'form': form, 'examination': examination})
        else:
            # Form is not valid (e.g., no file selected)
            # Pass examination context back to template if it was retrieved via GET
            context = {
                'form': form,
                'examination': examination # Pass the examination object back for display
            }
            for field, field_errors in form.errors.items():
                for error in field_errors:
                    messages.error(request, f"{field.replace('_', ' ').title()}: {error}")
            return render(request, 'students/mark_excel_upload.html', context)
    else:
        # GET request: Initial page load
        form = MarkExcelUploadForm()
        context = {
            'form': form,
            'examination': examination # Pass the examination object to the template
        }
        return render(request, 'students/mark_excel_upload.html', context)

@login_required
@user_passes_test(is_general_school_dashboard_user, login_url='/users/login/')
def mark_list(request):
    form = MarkEntrySelectionForm(request.GET or None)
    marks = Mark.objects.all()

    selected_examination = None
    selected_class = None
    selected_subject = None

    if form.is_valid():
        examination_filter = form.cleaned_data.get('examination')
        class_filter = form.cleaned_data.get('class_name')
        subject_filter = form.cleaned_data.get('subject')

        if examination_filter:
            marks = marks.filter(examination=examination_filter)
            selected_examination = examination_filter
        if class_filter:
            marks = marks.filter(student__current_class=class_filter)
            selected_class = class_filter
        if subject_filter:
            marks = marks.filter(subject=subject_filter)
            selected_subject = subject_filter

    if is_class_teacher(request.user):
        try:
            user_assigned_class = Class.objects.get(class_teacher=request.user)
            marks = marks.filter(student__current_class=user_assigned_class)
            if selected_class and selected_class.pk != user_assigned_class.pk:
                messages.warning(request, "As a Class Teacher, you can only view marks for your assigned class. Filter applied accordingly.")
            selected_class = user_assigned_class
        except Class.DoesNotExist:
            messages.error(request, "You are a Class Teacher but not assigned to any class. Please contact the administrator.")
            marks = Mark.objects.none() 
            form.fields['class_name'].queryset = Class.objects.none() 
            messages.info(request, "No class assigned. Please contact admin.")


    if not request.GET and not (selected_examination or selected_class or selected_subject):
        latest_exam = Examination.objects.order_by('-academic_year', '-term', '-date').first()
        if latest_exam:
            marks = marks.filter(examination=latest_exam)
            selected_examination = latest_exam
        else:
            marks = Mark.objects.none() 

    marks = marks.order_by(
        'examination__academic_year', 'examination__term',
        'student__current_class__name', 'subject__name', 'student__first_name'
    )

    context = {
        'form': form,
        'marks': marks,
        'selected_examination': selected_examination,
        'selected_class': selected_class,
        'selected_subject': selected_subject,
    }
    return render(request, 'students/mark_list.html', context)

@login_required
@user_passes_test(is_admin_or_teacher, login_url='/users/login/')
def result_selection(request):
    if request.method == 'POST':
        form = ResultSelectionForm(request.POST)
        if form.is_valid():
            examination = form.cleaned_data['examination']
            class_name = form.cleaned_data['class_name']
            query_params = f"?exam_id={examination.pk}"
            if class_name:
                query_params += f"&class_id={class_name.pk}"
                return redirect(f"{reverse('class_results_summary')}{query_params}")
            else:
                messages.error(request, "Please select a Class to view results.")
                return render(request, 'students/result_selection.html', {'form': form})
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field.replace('_', ' ').title()}: {error}")
    else:
        form = ResultSelectionForm()
        if is_class_teacher(request.user):
            try:
                assigned_class = Class.objects.get(class_teacher=request.user)
                form.fields['class_name'].queryset = Class.objects.filter(pk=assigned_class.pk)
                form.fields['class_name'].initial = assigned_class.pk
                form.fields['class_name'].widget.attrs['readonly'] = 'readonly'
                form.fields['class_name'].required = True 
            except Class.DoesNotExist:
                messages.warning(request, "You are a Class Teacher but not assigned to any class. Please contact the administrator.")
                form.fields['class_name'].queryset = Class.objects.none()
            except Class.MultipleObjectsReturned:
                messages.warning(request, "You are assigned to multiple classes. Please contact the administrator for data correction.")
                form.fields['class_name'].queryset = Class.objects.filter(class_teacher=request.user)

    return render(request, 'students/result_selection.html', {'form': form})

@login_required
@user_passes_test(is_admin_or_teacher, login_url='/users/login/')
def class_results_summary(request):
    exam_id = request.GET.get('exam_id')
    class_id = request.GET.get('class_id')

    if not exam_id or not class_id:
        messages.error(request, "Please select an examination and a class to view results.")
        return redirect('entry_selection')

    try:
        examination = Examination.objects.get(id=exam_id)
        class_obj = Class.objects.get(id=class_id)

        # --- IMPORTANT: Move the all_subjects query INSIDE this try block ---
        all_subjects = Subject.objects.filter(
            mark__examination=examination,
            mark__student__current_class=class_obj
        ).distinct().order_by('name')

    except (Examination.DoesNotExist, Class.DoesNotExist):
        messages.error(request, "Selected examination or class not found.")
        return redirect('entry_selection')

    # Now results and context can be created, as examination, class_obj, and all_subjects are guaranteed to exist here
    results = calculate_results(examination, class_obj)

    filtered_results = [
        r for r in results
        if r.get('subject_details') and any(sd.get('score') is not None for sd in r['subject_details'])
    ]


    context = {
        'examination': examination,
        'class_obj': class_obj,
        'results': filtered_results, 
        'all_subjects': all_subjects, # This variable is now guaranteed to be defined
    }
    return render(request, 'students/class_results_summary.html', context)

@login_required
@user_passes_test(is_admin_or_teacher, login_url='/users/login/')
def student_result_slip(request, exam_id, student_id):
    try:
        examination = get_object_or_404(Examination, pk=exam_id)
        student = get_object_or_404(Student, pk=student_id)
    except (Examination.DoesNotExist, Student.DoesNotExist):
        messages.error(request, "Invalid Examination or Student ID.")
        return redirect('result_selection')

    if is_class_teacher(request.user):
        try:
            user_assigned_class = Class.objects.get(class_teacher=request.user)
            if student.current_class.pk != user_assigned_class.pk:
                messages.error(request, "As a Class Teacher, you can only view result slips for students in your assigned class.")
                return redirect('result_selection')
        except Class.DoesNotExist:
            messages.error(request, "You are a Class Teacher but not assigned to any class. Please contact the administrator.")
            return redirect('result_selection')

    class_for_position = student.current_class
    all_results_in_class = calculate_results(examination, class_for_position)

    student_result = next((item for item in all_results_in_class if item['student'].id == student.id), None)

    if not student_result:
        messages.info(request, "No results found for this student in the selected examination.")
        return redirect('class_results_summary', exam_id=exam_id, class_id=student.current_class.pk) 

    context = {
        'examination': examination,
        'student': student,
        'student_result': student_result,
        'class_obj': class_for_position 
    }
    return render(request, 'students/student_result_slip.html', context)

@login_required
@user_passes_test(is_admin_or_teacher, login_url='/users/login/')
def attendance_view(request):
    context = {
        'message': 'Attendance management functionality coming soon!'
    }
    return render(request, 'students/attendance.html', context)

@login_required
@user_passes_test(is_admin_or_teacher, login_url='/users/login/')
def timetable_view(request):
    context = {
        'message': 'Timetable viewing functionality coming soon!'
    }
    return render(request, 'students/timetable.html', context)

@login_required
@user_passes_test(is_admin_or_headteacher_or_statistic_teacher) 
def add_student(request):
    if request.method == 'POST':
        form = StudentCreationForm(request.POST, user=request.user)
        if form.is_valid():
            student = form.save()
            messages.success(request, f"Student '{student.get_full_name}' (Admission No: {student.admission_number}) added successfully.")
            return redirect('student_list') 
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field.replace('_', ' ').title()}: {error}")
    else:
        form = StudentCreationForm(user=request.user)

    context = {
        'form': form,
        'title': 'Add New Student'
    }
    return render(request, 'students/student_form.html', context)

@login_required
@user_passes_test(is_admin_or_headteacher)
def edit_student(request, pk):
    student = get_object_or_404(Student, pk=pk)

    if is_class_teacher(request.user):
        try:
            user_assigned_class = Class.objects.get(class_teacher=request.user)
            if student.current_class.pk != user_assigned_class.pk:
                messages.error(request, "As a Class Teacher, you can only edit students in your assigned class.")
                return redirect('student_list')
        except Class.DoesNotExist:
            messages.error(request, "You are a Class Teacher but not assigned to any class. Please contact the administrator.")
            return redirect('student_list')

    if request.method == 'POST':
        form = StudentCreationForm(request.POST, instance=student, user=request.user)
        if form.is_valid():
            student = form.save()
            messages.success(request, f"Student '{student.get_full_name}' (Admission No: {student.admission_number}) updated successfully.")
            return redirect('student_list')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field.replace('_', ' ').title()}: {error}")
    else:
        form = StudentCreationForm(instance=student, user=request.user)

    context = {
        'form': form,
        'title': f'Edit Student: {student.get_full_name}'
    }
    return render(request, 'students/student_form.html', context)

@login_required
def home_view(request):
    if request.user.is_authenticated:
        if request.user.role == 'admin':
            return redirect(reverse('admin_dashboard'))
        # Explicitly list all roles that should go to the teacher_dashboard
        elif request.user.role in ['class_teacher', 'academic_teacher', 'headteacher', 'statistic_teacher', 'subject_teacher']:
            return redirect(reverse('teacher_dashboard'))
    return redirect(reverse('login'))

def process_student_excel_row(row_data, request):
    try:
        first_name = row_data.get('first_name')
        middle_name = row_data.get('middle_name')
        last_name = row_data.get('last_name')
        date_of_birth = row_data.get('date_of_birth')
        gender = row_data.get('gender')
        admission_number = row_data.get('admission_number')
        current_class_name = row_data.get('current_class_name')

        if not all([first_name, last_name, admission_number, current_class_name]):
            raise ValueError("Missing essential data (first name, last name, admission number, class name).")

        try:
            # Now 'Class' is defined because it's imported at the top
            current_class = Class.objects.get(name=current_class_name) 
        except Class.DoesNotExist:
            raise ValueError(f"Class '{current_class_name}' not found. Please ensure the class exists.")

        # Now 'Student' is defined because it's imported at the top
        student, created = Student.objects.update_or_create(
            admission_number=admission_number,
            defaults={
                'first_name': first_name,
                'middle_name': middle_name,
                'last_name': last_name,
                'date_of_birth': date_of_birth,
                'gender': gender,
                'current_class': current_class,
            }
        )
        return True, f"Student {student.get_full_name()} ({'created' if created else 'updated'})."
    except Exception as e:
        return False, f"Error: {e}"

@login_required
@user_passes_test(is_admin_or_headteacher_or_statistic_teacher, login_url='/users/login/') 
def upload_students_excel(request):
    if request.method == 'POST':
        form = StudentExcelUploadForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = form.cleaned_data['excel_file']

            if not excel_file.name.endswith(('.xlsx', '.xls')):
                messages.error(request, "Please upload a valid Excel file (.xlsx or .xls).")
                return redirect('upload_students_excel')

            try:
                df = pd.read_excel(excel_file)
                
                df.columns = df.columns.str.lower() 

                success_count = 0
                error_messages = []

                for index, row_data in df.iterrows():
                    row_number = index + 2 # Excel rows are 1-indexed, and header is row 1
                    success, message = process_student_excel_row(row_data.to_dict(), request)
                    if success:
                        success_count += 1
                    else:
                        error_messages.append(f"Row {row_number}: {message}")

                if success_count > 0:
                    messages.success(request, f"Successfully processed {success_count} student records.")
                if error_messages:
                    for msg in error_messages:
                        messages.warning(request, msg) # Use warning for row-level errors
                
                return redirect('upload_students_excel') # Or redirect to student list

            except ImportError:
                messages.error(request, "Please install pandas and openpyxl: pip install pandas openpyxl")
            except Exception as e:
                messages.error(request, f"An unexpected error occurred during file processing: {e}")
        else:
            messages.error(request, "Please correct the errors below.")
    else:
        form = StudentExcelUploadForm()
    
    context = {'form': form}
    return render(request, 'students/upload_students_excel.html', context)

@login_required
@user_passes_test(is_general_school_dashboard_user)
def teacher_dashboard(request):
    total_students_overall = Student.objects.count()
    total_boys_overall = Student.objects.filter(gender='M').count()
    total_girls_overall = Student.objects.filter(gender='F').count()

    # Initialize variables for the "My Class" card
    teacher_boys_count = 0
    teacher_girls_count = 0
    assigned_classes = Class.objects.none()

    if is_class_teacher(request.user):
        # For a Class Teacher, get stats for their assigned class(es)
        assigned_classes = Class.objects.filter(class_teacher=request.user)
        if assigned_classes.exists():
            students_in_teacher_classes = Student.objects.filter(current_class__in=assigned_classes)
            teacher_boys_count = students_in_teacher_classes.filter(gender='M').count()
            teacher_girls_count = students_in_teacher_classes.filter(gender='F').count()
        else:
            messages.info(request, "You are not currently assigned to a class.")

    elif is_academic_teacher(request.user):
        # For an Academic Teacher, the "My Class" card will show overall school stats
        teacher_boys_count = total_boys_overall
        teacher_girls_count = total_girls_overall

    context = {
        'total_boys': total_boys_overall,   
        'total_girls': total_girls_overall, 
        'teacher_class_boys': teacher_boys_count,
        'teacher_class_girls': teacher_girls_count,
        'assigned_classes': assigned_classes,
    }

    return render(request, 'users/teacher_dashboard.html', context)

@login_required
@user_passes_test(is_class_teacher) 
def students_in_my_class_view(request):
    students = Student.objects.none()
    message = None
    teacher_class_assignment = None
    examination = None 
    if hasattr(request.user, 'assigned_class') and request.user.assigned_class:
        teacher_class_assignment = request.user.assigned_class
        students = Student.objects.filter(current_class=teacher_class_assignment).order_by('first_name', 'last_name')

        # Try to get the latest examination for this class
        examination = Examination.objects.filter(classes_taking_exam=teacher_class_assignment).order_by('-academic_year', '-term').first()

        if not students.exists():
            message = f"No students are currently assigned to your class: {teacher_class_assignment.name}."
        if not examination and students.exists(): # Only show this message if there are students but no exam data
            message = (message or "") + (f" No examinations found for your class: {teacher_class_assignment.name}. Result slips and performance analysis may be incomplete.")

    else:
        message = "You are not assigned to any class or your profile is incomplete. Please contact the administrator."
    
    context = {
        'students': students,
        'page_title': 'Students in My Class',
        'message': message,
        'show_class_column': False, # This view is already specific to a class
        'class_obj': teacher_class_assignment, # Pass the class object for potential links
        'examination': examination, # Pass the examination object for potential links
    }
    return render(request, 'students/student_list_template.html', context)

@login_required
@staff_member_required
def all_students_view(request):
    students = Student.objects.all()
    page_title = 'All Students in School'
    show_class_column = True # Always show class column on this page

    # --- Data for Filter Dropdowns ---
    all_classes = Class.objects.all().order_by('name')
    # Get unique gender choices directly from the model field
    all_genders = Student.gender_choices
    # Get unique status choices directly from the model field
    all_statuses = Student.STATUS_CHOICES

    # --- Get Filter Parameters from GET request ---
    selected_class_id = request.GET.get('class_id', '') # Default to empty string
    selected_gender = request.GET.get('gender', '')     # Default to empty string
    selected_status = request.GET.get('status', '')     # Default to empty string
    search_query = request.GET.get('q', '')             # Default to empty string

    # --- Apply Filters to the Queryset ---

    # Filter by Class
    if selected_class_id:
        if selected_class_id == 'unassigned':
            students = students.filter(current_class__isnull=True)
            page_title += " (Unassigned Class)"
        else:
            try:
                selected_class_obj = Class.objects.get(id=selected_class_id)
                students = students.filter(current_class=selected_class_obj)
                page_title += f" ({selected_class_obj.name})"
            except Class.DoesNotExist:
                messages.warning(request, "Invalid class selected for filtering.")
                selected_class_id = '' # Reset to show no class selected in dropdown

    # Filter by Gender
    if selected_gender:
        students = students.filter(gender=selected_gender)
        # You might want to append to page_title based on gender too, e.g., " (Male)"

    # Filter by Status
    if selected_status:
        students = students.filter(status=selected_status)
        # You might want to append to page_title based on status too, e.g., " (Active)"

    # Search by Name or Admission Number
    if search_query:
        students = students.filter(
            Q(first_name__icontains=search_query) |
            Q(middle_name__icontains=search_query) | # <--- MIDDLE NAME ADDED HERE
            Q(last_name__icontains=search_query) |
            Q(admission_number__icontains=search_query)
        )
        page_title = f"Students matching '{search_query}'"


    # Order the results (can be combined with filtering above)
    students = students.order_by('current_class__name', 'first_name', 'last_name')

    context = {
        'students': students,
        'page_title': page_title,
        'show_class_column': show_class_column,
        'all_classes': all_classes,             # Pass all classes for dropdown
        'all_genders': all_genders,             # Pass all genders for dropdown
        'all_statuses': all_statuses,           # Pass all statuses for dropdown
        'selected_class_id': selected_class_id, # Pass back selected value to retain in dropdown
        'selected_gender': selected_gender,     # Pass back selected value to retain in dropdown
        'selected_status': selected_status,     # Pass back selected value to retain in dropdown
        'search_query': search_query,           # Pass back search query to retain in input
    }
    return render(request, 'students/student_list_template.html', context)

@staff_member_required 
def add_student_view(request):
    if request.method == 'POST':
        form = StudentForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Student added successfully!')
            return redirect('all_students')
        else:
            messages.error(request, 'Error adding student. Please check the form.')
    else:
        form = StudentForm()

    context = {
        'form': form,
        'page_title': 'Add New Student',
    }
    return render(request, 'students/add_student_form.html', context)

@login_required
@user_passes_test(is_admin_or_headteacher)
def delete_student_view(request, pk):
    student = get_object_or_404(Student, pk=pk)
    if request.method == 'POST':
        student.delete()
        messages.success(request, 'Student deleted successfully!')
        return redirect('all_students')
    # If it's a GET request, you might want to show a confirmation page
    context = {
        'student': student,
        'page_title': f'Confirm Delete: {student.first_name} {student.last_name}',
    }
    return render(request, 'students/confirm_delete_student.html', context)

def is_passing_grade(grade):
    """
    Returns True if the grade is considered passing, False otherwise.
    Adjust the logic as needed for your grading system.
    """
    return grade in ['A', 'B', 'C', 'D']

def get_grade_from_score(score):
    """
    Maps a numeric score to a grade character.
    """
    if score is None:
        return "N/A"
    if score >= 81:
        return "A"
    elif score >= 61:
        return "B"
    elif score >= 41:
        return "C"
    elif score >= 21:
        return "D"
    else:
        return "F"

@login_required
def performance_selection_view(request):
    classes = Class.objects.all().order_by('name')
    examinations = Examination.objects.all().order_by('-academic_year', '-term')

    # Filter classes if the user is a class_teacher
    if request.user.role == 'class_teacher' and request.user.assigned_class:
        classes = Class.objects.filter(pk=request.user.assigned_class.pk)
    
    # Handle form submission
    if request.method == 'POST':
        class_id = request.POST.get('class_id')
        exam_id = request.POST.get('exam_id')

        if class_id and exam_id:
            # Redirect to the actual performance analysis page
            return redirect('class_performance_analysis', class_id=class_id, exam_id=exam_id)
        else:
            # Handle error: user didn't select both
            return render(request, 'students/performance_selection.html', {
                'classes': classes,
                'examinations': examinations,
                'error_message': 'Please select both a Class and an Examination.',
                'page_title': 'Select Performance Criteria'
            })

    context = {
        'classes': classes,
        'examinations': examinations,
        'page_title': 'Select Performance Criteria'
    }
    return render(request, 'students/performance_selection.html', context)

def class_results_summary_view(request, class_id, exam_id):

    class_obj = get_object_or_404(Class, pk=class_id)
    examination = get_object_or_404(Examination, pk=exam_id)

    context = {
        'class_obj': class_obj,
        'examination': examination,
        'page_title': f'Class Results Summary - {class_obj.name}'
        # ... add your summary data to context ...
    }
    return render(request, 'students/class_results_summary.html', context)

@login_required
def class_performance_analysis_view(request, class_id, exam_id):
    class_obj = get_object_or_404(Class, pk=class_id)
    examination = get_object_or_404(Examination, pk=exam_id)

    # 1. Get all students in this class for this examination
    students_in_class_for_exam = Student.objects.filter(current_class=class_obj).distinct()
    total_students_in_class = students_in_class_for_exam.count()

    class_results = []
    for student in students_in_class_for_exam:
        marks = Mark.objects.filter(student=student, examination=examination)
        if marks.exists():
            total_score = sum(mark.score for mark in marks if mark.score is not None)
            num_subjects_scored = sum(1 for mark in marks if mark.score is not None)
            average_score = total_score / num_subjects_scored if num_subjects_scored > 0 else 0
            overall_grade = get_grade_from_score(average_score)

            class_results.append({
                'student': student,
                'total_score': total_score,
                'average_score': round(average_score, 2),
                'overall_grade': overall_grade,
                'is_pass': is_passing_grade(overall_grade)
            })

    class_results.sort(key=lambda x: x['total_score'], reverse=True)

    current_position = 1
    previous_score = None
    for i, result in enumerate(class_results):
        if result['total_score'] != previous_score:
            current_position = i + 1
        result['position'] = current_position
        previous_score = result['total_score']

    top_students = class_results[:10]

    overall_grade_distribution = {}
    for grade_char in ['A', 'B', 'C', 'D', 'E', 'F', 'N/A']:
        overall_grade_distribution[grade_char] = 0

    overall_pass_count = 0
    overall_fail_count = 0

    for result in class_results:
        overall_grade_distribution[result['overall_grade']] = overall_grade_distribution.get(result['overall_grade'], 0) + 1
        if result['is_pass']:
            overall_pass_count += 1
        else:
            overall_fail_count += 1

    overall_pass_rate = (overall_pass_count / total_students_in_class) * 100 if total_students_in_class > 0 else 0
    overall_fail_rate = (overall_fail_count / total_students_in_class) * 100 if total_students_in_class > 0 else 0  
    
    subject_analysis = []
    all_subjects = class_obj.subjects.all().order_by('name')

    for subject in all_subjects:
        subject_data = {
            'name': subject.name,
            'code': subject.code,
            'total_scored': 0,
            'grades': {'A': 0, 'B': 0, 'C': 0, 'D': 0, 'E': 0, 'F': 0},
            'pass_count': 0,
            'fail_count': 0,
        }

        marks_for_subject = Mark.objects.filter(
            examination=examination,
            subject=subject,
            student__current_class=class_obj
        ).select_related('student')

        for mark in marks_for_subject:
            if mark.score is not None:
                subject_data['total_scored'] += 1
                grade = get_grade_from_score(mark.score)
                subject_data['grades'][grade] = subject_data['grades'].get(grade, 0) + 1
                if is_passing_grade(grade):
                    subject_data['pass_count'] += 1
                else:
                    subject_data['fail_count'] += 1
                    
        if subject_data['total_scored'] > 0:
            subject_data['pass_percentage'] = (subject_data['pass_count'] / subject_data['total_scored']) * 100
            subject_data['fail_percentage'] = (subject_data['fail_count'] / subject_data['total_scored']) * 100
        else:
            subject_data['pass_percentage'] = 0 # Ensure 0 if no students scored
            subject_data['fail_percentage'] = 0 # Ensure 0 if no students scored
        
        subject_analysis.append(subject_data)

    context = {
        'examination': examination,
        'class_obj': class_obj,
        'total_students_in_class': total_students_in_class,
        'overall_grade_distribution': overall_grade_distribution,
        'overall_pass_count': overall_pass_count,
        'overall_fail_count': overall_fail_count,
        'overall_pass_rate': overall_pass_rate,
        'overall_fail_rate': round(overall_fail_rate, 2),
        'overall_pass_rate': round(overall_pass_rate, 2),
        'subject_analysis': subject_analysis,
        'top_students': top_students,
        'page_title': f'Class Performance - {class_obj.name}'
    }
    return render(request, 'students/class_performance_analysis.html', context)

@login_required
def student_result_slip_view(request, student_id, examination_id):
    student = get_object_or_404(Student, pk=student_id)
    examination = get_object_or_404(Examination, pk=examination_id)

    marks = Mark.objects.filter(student=student, examination=examination)

    total_score = 0
    subject_details = []
    num_subjects_scored = 0

    for mark in marks:
        if mark.score is not None:
            total_score += mark.score
            num_subjects_scored += 1
            subject_details.append({
                'subject_name': mark.subject.name,
                'score': mark.score,
                'grade': get_grade(mark.score)
            })
    
    average_score = (total_score / num_subjects_scored) if num_subjects_scored > 0 else 0
    overall_grade = get_grade_from_score(average_score)

    student_position = None # Or implement logic to calculate class position here

    student_result = {
        'total_score': total_score,
        'average_score': round(average_score, 2),
        'overall_grade': overall_grade,
        'subject_details': subject_details,
        'position': student_position,
    }

    context = {
        'student': student,
        'examination': examination,
        'student_result': student_result,
    }
    return render(request, 'students/student_result_slip.html', context)

class AdminRequiredMixin(UserPassesTestMixin):
    def test_func(self):
        return self.request.user.is_staff

# List all documents
class DocumentListView(LoginRequiredMixin, AdminRequiredMixin, ListView):
    model = SchoolDocument
    template_name = 'students/document_list.html'
    context_object_name = 'documents'
    ordering = ['-uploaded_at']

# Upload a new document
class DocumentUploadView(LoginRequiredMixin, AdminRequiredMixin, CreateView):
    model = SchoolDocument
    form_class = SchoolDocumentForm
    template_name = 'students/document_upload_form.html'
    success_url = reverse_lazy('document_list')

    def form_valid(self, form):
        form.instance.uploaded_by = self.request.user # Assuming the logged-in user is the uploader
        return super().form_valid(form)

def calculate_student_result(student, examination):
    """
    Helper function to calculate all results for a student in a given exam.
    This version uses your CustomUser model and the 'role' field.
    """
    # ... (all your existing calculation logic for position, total, average, etc.) ...
    students_in_class = Student.objects.filter(current_class=student.current_class)

    class_results = []
    for s in students_in_class:
        marks = Mark.objects.filter(student=s, examination=examination)
        total = sum(mark.score or 0 for mark in marks)
        class_results.append({
            'student_id': s.id,
            'total_score': total,
            'marks_count': marks.count(),
        })

    sorted_results = sorted(class_results, key=lambda x: x['total_score'], reverse=True)
    for i in range(len(sorted_results)):
        if i > 0 and sorted_results[i]['total_score'] == sorted_results[i-1]['total_score']:
            sorted_results[i]['position'] = sorted_results[i-1]['position']
        else:
            sorted_results[i]['position'] = i + 1

    student_result_data = next((res for res in sorted_results if res['student_id'] == student.id), None)
    student_marks = Mark.objects.filter(student=student, examination=examination)
    subject_details = []
    for mark in student_marks:
        subject_details.append({
            'subject_name': mark.subject.name,
            'score': mark.score,
            'grade': get_grade(mark.score)
        })

    total_score = student_result_data['total_score']
    average_score = Decimal(total_score) / Decimal(student_result_data['marks_count']) if student_result_data['marks_count'] > 0 else None
    overall_grade = get_grade(average_score)

    # NEW: Fetch the class teacher and head teacher using your model structure
    class_teacher = student.current_class.class_teacher
    head_teacher = CustomUser.objects.filter(role='headteacher').first()
    
    return {
        'position': student_result_data['position'],
        'total_score': total_score,
        'average_score': average_score,
        'overall_grade': overall_grade,
        'subject_details': subject_details,
        'class_teacher': class_teacher,
        'head_teacher': head_teacher,
    }

def download_class_summary_pdf(request, exam_id, class_id):
    examination = get_object_or_404(Examination, pk=exam_id)
    class_obj = get_object_or_404(Class, pk=class_id)
    
    # ... (your data retrieval and calculation logic remains the same) ...
    # This part of the code is unchanged from my last response
    all_subjects = Subject.objects.filter(classes_assigned=class_obj).order_by('code')
    students_in_class = Student.objects.filter(current_class=class_obj).order_by('admission_number')
    
    results = []
    
    for student in students_in_class:
        marks = Mark.objects.filter(student=student, examination=examination)
        total_score = sum(mark.score or 0 for mark in marks)
        
        if marks.exists():
            average_score = Decimal(total_score) / Decimal(marks.count())
        else:
            average_score = None
            
        overall_grade = get_grade(average_score)

        subject_details = [{'subject_code': mark.subject.code, 'score': mark.score} for mark in marks]
        
        results.append({
            'student': student,
            'total_score': total_score,
            'average_score': average_score,
            'overall_grade': overall_grade,
            'subject_details': subject_details,
            'position': None
        })

    sorted_results = sorted(results, key=lambda x: x['total_score'], reverse=True)

    if sorted_results:
        sorted_results[0]['position'] = 1
        for i in range(1, len(sorted_results)):
            if sorted_results[i]['total_score'] == sorted_results[i-1]['total_score']:
                sorted_results[i]['position'] = sorted_results[i-1]['position']
            else:
                sorted_results[i]['position'] = i + 1

    # 2. Render HTML and generate PDF
    html_string = render_to_string('students/class_results_pdf.html', {
        'examination': examination,
        'class_obj': class_obj,
        'results': sorted_results,
        'all_subjects': all_subjects
    })

    # 3. Use a query parameter to determine 'inline' vs 'attachment'
    response = HttpResponse(content_type='application/pdf')
    filename = f"Class_{class_obj.name}_{examination.name}_Summary.pdf"

    # Check the URL for the 'download' parameter
    if request.GET.get('download', 'false').lower() == 'true':
        # This will trigger a file download
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
    else:
        # This will display the PDF directly in the browser for printing
        response['Content-Disposition'] = f'inline; filename="{filename}"'

    HTML(string=html_string).write_pdf(response)

    return response

def download_student_result_pdf(request, exam_id, student_id):
    examination = get_object_or_404(Examination, pk=exam_id)
    student = get_object_or_404(Student, pk=student_id)
    
    # Use the helper function to get the correct data
    student_result = calculate_student_result(student, examination)

    current_year = date.today().year

    context = {
        'examination': examination,
        'student': student,
        'student_result': student_result,
        'current_year': current_year,
    }

    html_string = render_to_string('students/student_result_pdf.html', context)
    
    response = HttpResponse(content_type='application/pdf')
    filename = f"Result_Slip_{student.get_full_name().replace(' ', '_')}_{examination.academic_year}.pdf"

    if request.GET.get('download', 'false').lower() == 'true':
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
    else:
        response['Content-Disposition'] = f'inline; filename="{filename}"'
        
    HTML(string=html_string).write_pdf(response)
    
    return response

def view_student_result_slip(request, exam_id, student_id):
    examination = get_object_or_404(Examination, pk=exam_id)
    student = get_object_or_404(Student, pk=student_id)
    
    # Use the helper function to get the correct, consistent data
    student_result = calculate_student_result(student, examination)

    context = {
        'examination': examination,
        'student': student,
        'student_result': student_result,
    }
    return render(request, 'students/result_slip.html', context)
